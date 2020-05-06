print('monitor')
import monitor
print('threading')
import threading
print('datetime')
import datetime
print('soupRequests')
from soupRequests import RequestBot
print('openpyxl')
import openpyxl
print('time')
import time
print('win10toast')
from win10toast import ToastNotifier

user_end_command = False

toaster = 2
def setToaster():
	global toaster
	print('hello')
	toaster = ToastNotifier()
	toaster.show_toast("Coronavirus", "5G Tower has activated.", threaded=True, icon_path=None, duration=2)
	print('ugh')

x = threading.Thread(target = setToaster)
x.start()
x.join()

path = 'C:\\Users\\isaac\\Desktop\\desktop python 4-27-20\\eee_coronavirus.xlsx'
wb = openpyxl.load_workbook(path) 
ws = wb['World']
while user_end_command == False:
	# pull last row data
	empty = False
	empty_row = 2
	while not empty:
		if ws.cell(row = empty_row, column = 1).value == None or ws.cell(row = empty_row, column = 1).value == '':
			empty = True
		else:
			empty_row += 1

	above_row = empty_row - 1
	last_value = ws.cell(row = above_row, column = 2).value

	# pulling website data
	try:
		world = RequestBot('https://www.worldometers.info/coronavirus/')
		new_value = world.soup.find('span', style = 'color:#aaa').text
	except:
		toaster.show_toast("Coronavirus", "Cannot establish connection to 5G Tower, retrying in 10 seconds", threaded=True, icon_path=None, duration=6)
		print('no wifi')
		time.sleep(10)
	else:

		if last_value == new_value:
			print(f'{last_value} is the same as {new_value}')
			time.sleep(5)
		else:
			print(f'{last_value} is not the same as {new_value}')
			x = threading.Thread(target = monitor.main)
			x.start()
			x.join()
			toaster.show_toast("Coronavirus", "The coronavirus 5G Tower has updated!", threaded=True, icon_path=None, duration=6)
