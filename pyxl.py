	from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Alignment

class ExcelMaker(object):
	def __init__(self, headers):
		self.headers = list(headers)
	def columnWidth(self, width):
		import string
		alpha_dict = dict(enumerate(string.ascii_lowercase))
		for i in range(0, len(self.headers)): 
			letter = str(alpha_dict[i])
			self.ws.column_dimensions[letter].width = width
	def createHeaders(self):
		for header in self.headers:
			col = self.headers.index(header) + 1
			self.ws.cell(row = 1, column = col, value = header)
	def loadWorkbook(self, path, wsName):
		self.wb = openpyxl.load_workbook(path)
		try:
			self.ws = self.wb[wsName]
			self.createHeaders()
		except KeyError: # meaning there is no worksheet called "wsName"
			self.ws = self.wb.create_sheet(wsName)
			self.createHeaders()
			print(f'no worksheet called "{wsName}", creating new worksheet')
	def newWorkbook(self, wsName):
		self.wb = openpyxl.Workbook() 
		self.ws = wb.active
		self.ws.title = wsName
		self.createHeaders()
		

	def findNewRow(self):
		self.empty = False
		self.row = self.first_row = 2 # this number should be the first row after the headings
		while not self.empty:
			value = self.ws.cell(row = self.row, column = 1).value
			if value != None:
				# print(f'{value} does not equal {None}')
			#if ws['A' + str(self.row)].value != None:
				self.value = self.ws['A' + str(self.row)].value
				self.row += 1
				self.empty = False
			else:
				self.empty = True
	def sameDataCheck(self,firstColData):
		value = self.ws.cell(row = self.row - 1, column = 1).value
		if value == firstColData:
			return True #, they are the same data, do not add more of the same data
		else:
			return False
	def addData(self, dataList):
		for data in dataList:
			self.ws.cell(row = self.row, column = dataList.index(data) + 1, value = data)
			self.ws.cell(row = self.row, column = dataList.index(data)).alignment = Alignment(horizontal = 'center')
	def close(self, path):
		try:
			self.wb.save(path)
		except PermissionError:
			print('Cannot save to excel file because the file is open. Please close the excel file and re-run this program.')
if __name__ == '__main__':
	pass