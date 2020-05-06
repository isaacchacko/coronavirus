print('initalizing "threading"')
import threading
print('initalizing "datetime"')
import datetime
print('initalizing "soupRequests"')
from soupRequests import RequestBot
print('initalizing "openpyxl"')
import openpyxl
print('initalizing "time"')
import time
print('initalizing "win10toast"')
from win10toast import ToastNotifier
print('initalizing "append"')
import append
def clean(text):
	text = text.replace(',','')
	text = text.strip()
	try:
		text = int(text)
	except:
		pass
	return text
class Runner(object):
	def __init__(self, path):
		self.path = path
	def run(self):
		print('Fetching "World"')
		self.world = Location('https://www.worldometers.info/coronavirus/#countries','World', self.path)
		print('Fetching "USA"')
		self.usa = Location('https://www.worldometers.info/coronavirus/country/us/','USA', self.path)
		print('Fetching "Spain"')
		self.spain = Location('https://www.worldometers.info/coronavirus/country/spain/','Spain', self.path)
		print('Fetching "Italy"')
		self.italy = Location('https://www.worldometers.info/coronavirus/country/italy/','Italy', self.path)
		print('Fetching "France"')
		self.france = Location('https://www.worldometers.info/coronavirus/country/france/','France', self.path)
		# uk = Location('https://www.worldometers.info/coronavirus/country/uk/','UK')
		print('Fetching "Germany"')
		self.germany = Location('https://www.worldometers.info/coronavirus/country/germany/','Germany', self.path)
		# turkey = Location('https://www.worldometers.info/coronavirus/country/turkey/','Turkey')
		# russia = Location('https://www.worldometers.info/coronavirus/country/russia/','Russia')
		print('Fetching "Iran"')
		self.iran = Location('https://www.worldometers.info/coronavirus/country/iran/','Iran', self.path)


class Location(object):
	def __init__(self, url, title, path):
		self.page = RequestBot(url)
		self.soup = self.page.soup
		self.title = title
		self.path = path
		self.main()
	def main(self):
		self.total_cases = clean(self.soup.find('span', style = 'color:#aaa').text)

		self.total_infected = self.soup.findAll('div', class_ = 'number-table-main')
		if len(self.total_infected) == 2:
			self.finished_cases = clean(self.total_infected[1].text)
			self.total_infected = clean(self.total_infected[0].text)
			self.weird = False
		else:
			self.weird = True
			self.finished_cases = clean(self.total_infected[0].text)
			self.total_infected = ''

		self.mild_condition = self.soup.findAll('span', class_ = 'number-table')
		if len(self.mild_condition) == 4:	
			self.serious_condition = clean(self.mild_condition[1].text)
			self.discharged = clean(self.mild_condition[2].text)
			self.total_dead = clean(self.mild_condition[3].text)
			self.mild_condition = clean(self.mild_condition[0].text)
		else:
			self.discharged = clean(self.mild_condition[0].text)
			self.total_dead = clean(self.mild_condition[1].text)
			self.serious_condition, self.mild_condition = '', ''


		divisor = float(self.total_cases) / 100.0
		self.death_rate = float(self.total_dead) / divisor 
		self.death_rate = self.death_rate * 0.01
		fmt = '%m/%d/%y %I:%M %p'
		self.time = time.strftime(fmt)
		if self.weird:
			x = threading.Thread(target = append.appendToXL, args = [['Time','Total Infected (All Time)', 'Total Finished Cases','Total Deaths','Total Discharges', 'Death Rate'],
				[self.time,
				self.total_cases,
				self.finished_cases,
				self.total_dead,
				self.discharged,
				self.death_rate],
				self.path, 
				self.title])
			x.start()
			x.join()
		else:
			x = threading.Thread(target = append.appendToXL, args = [['Time','Total Infected (All Time)', 'Total Finished Cases','Total Deaths','Total Discharges','Total Active Cases','Total Mild Condition','Total Serious Condition', 'Death Rate'],
				[self.time,
				self.total_cases,
				self.finished_cases,
				self.total_dead,
				self.discharged,
				self.total_infected,
				self.mild_condition,
				self.serious_condition,
				self.death_rate],
				self.path, 
				self.title])
			x.start()
			x.join()






class Infinity(object):
	def __init__(self, path):
		print('initalizing Toaster')
		self.startToaster()
		print('Toaster initalized')
		self.excel_file_path = path
		print(f'Excel File Path: {path}')
		self.excel_sheet_name = 'World'
		self.webpage_path = 'https://www.worldometers.info/coronavirus/'

		print('loaded excel file sucessfully')


		self.runner = Runner(path = self.excel_file_path)
	def startToaster(self):
		def start():
			self.toaster = ToastNotifier()
			self.notify('Coronavirus monitoring in progress')
		x = threading.Thread(target = start)
		x.start()
		x.join()
	def notify(self, msg, length = 4):
		self.toaster.show_toast("Infinity", msg, threaded=True, icon_path=None, duration=length)
	def mainloop(self):
		self.count = 0
		kill_switch = False
		while not kill_switch:
			self.last_value = self.findLastValue()
			try:
				self.new_value = self.findNewestValue()
			except:
				self.noInternet()
			else:
				if self.last_value != self.new_value:
					self.count += 1
					x = threading.Thread(target = self.runner.run)
					x.start()
					x.join()
					self.notify(f'New data has been found.\nCount: {self.count}')
				else:
					print(f'...too quiet... Count: {self.count}')
					time.sleep(5)
	def findLastValue(self):
		# pull last row data
		self.wb = openpyxl.load_workbook(self.excel_file_path) 
		self.ws = self.wb['World']
		empty = False
		empty_row = 2 # 1 is the header
		while not empty:
			if self.ws.cell(row = empty_row, column = 1).value == None or self.ws.cell(row = empty_row, column = 1).value == '':
				empty = True
			else:
				empty_row += 1

		above_row = empty_row - 1
		return self.ws.cell(row = above_row, column = 2).value
	def findNewestValue(self):
		x = RequestBot('https://www.worldometers.info/coronavirus/')
		return clean(x.soup.find('span', style = 'color:#aaa').text)
	def noInternet(self):
		self.notify('No Internet Connection Found. Retrying in 10 seconds')
		time.sleep(10)

if __name__ == '__main__':
	finn = Infinity('coronavirus_statistics.xlsx')
	finn.mainloop()