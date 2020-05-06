import openpyxl
import time
def appendToXL(headers, data_tuplee, path, title):
	wb = openpyxl.load_workbook(path)
	if 'Sheet1' in wb.sheetnames:
		wb.remove_sheet(wb['Sheet1'])
	try:
		ws = wb[title]
	except:
		ws = wb.create_sheet(title)

	# creates the headers
	for header in headers:
		col = headers.index(header) + 1
		ws.cell(row = 1, column = col, value = header)

	# find new row
	empty = False
	empty_row = 2
	while not empty:
		if ws.cell(row = empty_row, column = 1).value == None or ws.cell(row = empty_row, column = 1).value == '':
			empty = True
		else:
			empty_row += 1
	# print(f'empty row is {empty_row}')
	# checks if its a repeat data
	above_row = empty_row - 1
	if ws.cell(row = above_row, column = 2).value == data_tuplee[1]:
		# print(f'this is being tripped accidentally, {ws.cell(row = above_row, column = 2).value} equals {data_tuplee[1]}')
		print('no new data found')
	else:
		# appends the data
		for data in data_tuplee:
			col = data_tuplee.index(data) + 1
			ws.cell(row = empty_row, column = col, value = data)


	# saves the data
	try:
		wb.save(path)
	except PermissionError:
		input('Cannot save to excel file because the file is open. Please close the excel file and re-run this program.')